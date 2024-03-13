# -*- coding: utf-8 -*-
from konlpy.tag import Okt
from tqdm import tqdm
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.cluster import DBSCAN
import numpy as np


from bs4 import BeautifulSoup
from datetime import timedelta, datetime
import requests
import pandas as pd
import re

import openpyxl
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
import openpyxl.utils.cell
import time
import os



def makedir(dir):
    try:
        if not os.path.exists(dir):
            os.makedirs(dir)
    except OSError:
        print("Error: Failed to create the directory.")




#엑셀로 저장하기 위한 변수
RESULT_PATH = os.getcwd() + "/" + "crawling_result/"  #결과 저장할 경로
makedir(RESULT_PATH)

now = datetime.now() #파일이름 현 시간으로 저장하기

today_str = datetime.today().strftime("%Y%m%d")
today_num = int(today_str)


#날짜 정제화 함수
def date_cleansing(test):
    try:
        #지난 뉴스
        #머니투데이  10면1단  2018.11.05.  네이버뉴스   보내기
        pattern = '\d+.(\d+).(\d+).'  #정규표현식

        r = re.compile(pattern)
        match = r.search(test).group(0)  # 2018.11.05.
        return match

    except AttributeError:
        #최근 뉴스
        #이데일리  1시간 전  네이버뉴스   보내기
        pattern = '\w* (\d\w*)'     #정규표현식

        r = re.compile(pattern)
        match = r.search(test).group(1)
        #print(match)
        return match


#내용 정제화 함수
def contents_cleansing(contents):
    first_cleansing_contents = re.sub('<dl>.*?</a> </div> </dd> <dd>', '',
                                      str(contents)).strip()  #앞에 필요없는 부분 제거
    second_cleansing_contents = re.sub('<ul class="relation_lst">.*?</dd>', '',
                                       first_cleansing_contents).strip()#뒤에 필요없는 부분 제거 (새끼 기사)
    third_cleansing_contents = re.sub('<.+?>', '', second_cleansing_contents).strip()
    return third_cleansing_contents



def crawler(maxpage, query, sort, s_date, e_date,news_keyword):
    # 각 크롤링 결과 저장하기 위한 리스트 선언
    Crawl_date   = []
    News_keyword = []
    News_dates   = []
    Press_name   = []
    Title        = []
    Link         = []
    Contents     = []
    Page         = []

    result = {}

    s_from = re.sub(r'[^0-9]', '', s_date)
    e_to   = re.sub(r'[^0-9]', '', e_date)

    page = 1  # 한페이지당 10개 기사
    maxpage_t =(int(maxpage)-1)*10+1   # 11= 2페이지 21=3페이지 31=4페이지  ...81=9페이지 , 91=10페이지, 101=11페이지

    while page <= maxpage_t:
        url = "https://search.naver.com/search.naver?where=news&query=" + query + "&sort="+sort+"&ds=" + s_date + "&de=" + e_date + "&nso=so%3Ar%2Cp%3Afrom" + s_from + "to" + e_to + "%2Ca%3A&start=" + str(page)
        print(url)
        response = requests.get(url)
        time.sleep(1)  # url 웹 페이지 로딩 대기 시간

        html = response.text

        #뷰티풀소프의 인자값 지정
        soup = BeautifulSoup(html, 'html.parser')

        for news_result in soup.select(".list_news > li"):

            news_dates = news_result.select('span.info')
            press_name = news_result.select_one(".info.press").text
            title      = news_result.select_one(".news_tit").text
            link       = news_result.select_one(".news_tit")["href"]
            contents   = news_result.select_one(".news_dsc").text

            # (A면1단, 날짜) 형태로 되어 있는 경우 처리
            if len(news_dates) > 1:
                news_date = news_dates[1].get_text().replace('.', '')
            else:
                news_date = news_dates[0].get_text().replace('.', '')

            if "분" in news_date or "시간 전" in news_date:
                news_date = str(today_num)
                news_date = datetime.strptime(news_date, '%Y%m%d').strftime('%Y-%m-%d')

            elif "일 전" in news_date:
                news_date_num = int(re.sub('[\D]', '', news_date))
                news_date = str(datetime.today() - timedelta(days=news_date_num))
                news_date = news_date[:10]
            # news_date = datetime.strptime(news_date, '%Y%m%d').strftime('%Y-%m-%d')

            elif "주 전" in news_date:
                news_date_num = int(re.sub('[\D]', '', news_date))
                news_date = str(datetime.today() - timedelta(days=news_date_num * 7))
                news_date = news_date[:10]
            # news_date = datetime.strptime(news_date, '%Y%m%d').strftime('%Y-%m-%d')

            else:
                news_date = str(news_date)
                news_date = datetime.strptime(news_date, '%Y%m%d').strftime('%Y-%m-%d')

            Crawl_date.append(today_str)                            # 수집일자
            News_keyword.append(news_keyword)                       # 검색어
            News_dates.append(re.sub(r'[^0-9]', '', news_date))     # 기사일자
            Press_name.append(press_name)                           # 언론사
            Title.append(title)                                     # 기사제목
            Link.append(link)                                       # 기사URL
            Contents.append(contents_cleansing(contents))
            Page.append(page)

        print(page)
        page += 10


    #모든 리스트 딕셔너리형태로 저장
    result= {
             "수집일자"  : Crawl_date,       #A
             "검색어"    : News_keyword,     #B
             "기사일자"  : News_dates ,      #C
             "언론사"    : Press_name ,      #D
             "기사제목"  : Title ,           #E
             "기사링크"  : Link ,            #F
             "본문요약"  : Contents,         #G
             "페이지"    : Page              #H
             }

    df = pd.DataFrame(result)  # df로 변환

    print(df['기사제목'])

    okt = Okt()  # 형태소 분석기 객체 생성
    noun_list = []
    for content in tqdm(df['기사제목']):
        nouns = okt.nouns(content)  # 명사만 추출하기, 결과값은 명사 리스트
        noun_list.append(nouns)

    print(noun_list)
    df["nouns"] = noun_list
    print(df)

    # 명사가 비어있으면 제거
    drop_index_list = []  # 지워버릴 index를 담는 리스트
    for i, row in df.iterrows():
        temp_nouns = row['nouns']
        if len(temp_nouns) == 0:  # 만약 명사리스트가 비어 있다면
            drop_index_list.append(i)  # 지울 index 추가

    df = df.drop(drop_index_list)  # 해당 index를 지우기

    # index를 지우면 순회시 index 값이 중간중간 비기 때문에 index를 다시 지정
    df.index = range(len(df))
    print(df)

    # 문서를 명사 집합으로 보고 문서 리스트로 치환 (tfidfVectorizer 인풋 형태를 맞추기 위해)
    text = [" ".join(noun) for noun in df['nouns']]

    tfidf_vectorizer = TfidfVectorizer(min_df=5, ngram_range=(1, 5))
    tfidf_vectorizer.fit(text)
    vector = tfidf_vectorizer.transform(text).toarray()

    print(vector)

    vector = np.array(vector)  # Normalizer를 이용해 변환된 벡터
    model = DBSCAN(eps=0.3, min_samples=6, metric="cosine")
    # 거리 계산 식으로는 Cosine distance를 이용
    result = model.fit_predict(vector)

    print(result)
    df["result"] = result

    print(df)

    for cluster_num in set(result):
        # -1,0은 노이즈 판별이 났거나 클러스터링이 안된 경우
        if (cluster_num == -1 or cluster_num == 0):
            continue
        else:
            print("cluster num : {}".format(cluster_num))
            temp_df = df[df['result'] == cluster_num]  # cluster num 별로 조회
            for title in temp_df['기사제목']:
                print(title)  # 제목으로 살펴보자
            print()






    # 새로 만들 파일이름 지정
    resultFileName = 'RESULT_%04d%02d%02d_%02d%02d%02d_%s.xlsx' % (now.year, now.month, now.day, now.hour, now.minute, now.second,news_keyword)

    df.to_excel(
                 RESULT_PATH + resultFileName     # 파일명
               , sheet_name  = news_keyword       # 검색어
               , index       = False              # 인덱스
               , freeze_panes= (1,0)              # 틀고정
               )



    # 엑셀 스타일
    wb = load_workbook(RESULT_PATH + resultFileName)
    ws = wb.active

    ws.column_dimensions["A"].width = 10.25    # 수집일자
    ws.column_dimensions["B"].width = 10.25    # 검색어
    ws.column_dimensions["C"].width = 10.25    # 기사일자
    ws.column_dimensions["D"].width = 15       # 언론사
    ws.column_dimensions["E"].width = 50       # 기사제목
    ws.column_dimensions["F"].width = 50       # 기사링크
    ws.column_dimensions["G"].width = 100      # 본문요약
    ws.column_dimensions["H"].width = 8        # 페이지

    # 폰트 정의
    font_header  = Font(name="맑은 고딕", size=9, bold=True)
    font_content = Font(name="맑은 고딕", size=9, bold=False)

    cell_alignment_center = Alignment(horizontal    = 'center',  # 좌우 정렬 left, right, center, distributed
                                      vertical      = 'center',  # 위아래 정렬 top, bottom, center, distributed
                                      shrink_to_fit = True,      # 셀의 크기에 맞게 글자를 축소
                                      indent        = 0          # 들여쓰기
                              )

    # CELL 음영 정의
    ligthGrayFill = PatternFill(start_color='00C0C0C0',end_color  ='00C0C0C0',fill_type  ='solid')

    # 헤더CELL 스타일 적용
    for rows in ws["A1":"H1"]:
        for cell in rows:
            cell.font      = font_header
            cell.alignment = cell_alignment_center
            cell.fill      = ligthGrayFill

    # 내용CELL 스타일 적용
    for rows in ws["A2":"H1000"]:
        for cell in rows:
            cell.font      = font_content
            if cell.column_letter in ('A','C') :
                cell.alignment = cell_alignment_center






            # 엑셀 스타일 적용 파일 저장
    wb.save(RESULT_PATH + resultFileName)



def main():
    # maxpage = "1"  #input("최대 크롤링할 페이지 수 입력하시오: ")
    # query = "BNK" #input("검색어 입력: ")
    # sort = "0"   # input("뉴스 검색 방식 입력(관련도순=0  최신순=1  오래된순=2): ")    #관련도순=0  최신순=1  오래된순=2
    # s_date = "2024/03/01" # input("시작날짜 입력(2019.01.04):")  #2019.01.04
    # e_date = "2024/03/07" # input("끝날짜 입력(2019.01.05):")   #2019.01.05

    # 엑셀 파일 열기
    wb = openpyxl.load_workbook(filename = os.getcwd() + "/" + "ors_news_crawler.xlsx")

    # 시트 선택하기
    option = wb.get_sheet_by_name('OPTION')

    s_date  = option["A3"].value # 검색 시작일자 YYYY.MM.DD
    e_date  = option["B3"].value # 검색 종료일자 YYYY.MM.DD
    maxpage = option["C3"].value # 검색 page (1page = 기사 10건)
    sort    = option["D3"].value # 검색 옵션(정렬) 관련도순:0, 최신순:1, 오래된순:2

    news_keyword   = ""          # 뉴스 검색 키워드
    query          = ""          # 쿼리 string

    flag = False
    keyword_range = option["A5":"D19"]
    print(keyword_range)
    for keyword in keyword_range:
        query = ""  # 검색어 초기화

        for cell in keyword:
            if cell.column_letter == "A" and cell.value == None:    # 수집포함 메인 keyword None
                flag = True
                break      # 수집중단

            if cell.column_letter == "A" and cell.value != None:  # 수집포함 메인 keyword
                news_keyword = cell.value

            # print("cell_letter:" + str(cell.column_letter))
            # print("keyword:" + str(cell.value))

            if cell.column_letter in ('A','B') and cell.value != None:
                query += " " + "+" + str(cell.value)   # +검색어 생성

            if cell.column_letter in ('C','D') and cell.value != None:
                query += " " + "-" + str(cell.value)   # -검색어 생성

        if flag:
            break #수집 중단

        print("query" + query)
        crawler(maxpage, query, sort, s_date, e_date,news_keyword)
main()













