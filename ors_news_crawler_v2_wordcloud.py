# -*- coding: utf-8 -*-

import urllib.request
from urllib.parse import urlparse, parse_qs

import konlpy
from konlpy.tag import Okt
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.cluster import DBSCAN

import numpy as np


from bs4 import BeautifulSoup
from datetime import timedelta, datetime
import requests
import pandas as pd
import re

import openpyxl
import openpyxl.utils.cell
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.drawing.image import Image  as Image_openpyxl

from collections import Counter

from wordcloud import WordCloud
from wordcloud import ImageColorGenerator
from PIL import Image # 만약 "No module named 'PIL'" 에러가 발생하면 [ pip install Pillow==5.4.1 ] 로 라이브러리를 설치
import matplotlib.pyplot as plt


import time
import os
import sys



########################################################################################################################
def makedir(dir):
    try:
        if not os.path.exists(dir):
            os.makedirs(dir)
    except OSError:
        print("Error: Failed to create the directory.")


#프로그램 실행파일 경로
ROOT_PATH = os.getcwd()
print("WORK_ROOT : " + ROOT_PATH )

#엑셀로 저장하기 위한 경로
RESULT_PATH = ROOT_PATH + "\\" + "crawling_result"   #결과 저장할 경로
print("RESULT_PATH : " + RESULT_PATH)

makedir(RESULT_PATH)

now = datetime.now() #파일이름 현 시간으로 저장하기

today_str = datetime.today().strftime("%Y%m%d")
today_num = int(today_str)

# Pandas 출력 최대개수
pd.options.display.max_rows    = 60    #None
pd.options.display.max_columns = 30    #None


# 언론사(PRESS) ID 수집 - 딕셔너리 정의
PRESS_DIC ={}
url_press = 'https://news.naver.com/main/officeList.naver'
html_press = urllib.request.urlopen(url_press).read()
soup_press = BeautifulSoup(html_press,'html.parser')
press_list = soup_press.find_all(class_='list_press nclicks(\'rig.renws2pname\')')

for press in press_list :
    parts = urlparse(press.attrs['href'])

    press_name = press.get_text().strip()
    press_id   = parse_qs(parts.query)['officeId'][0]
    #print(press_name + ' : ' + press_id)

    PRESS_DIC[press_name] = press_id
print(PRESS_DIC)

########################################################################################################################




def cluster_cleansing(df):
    # 불필요 기사 cleansing - 명사가 비어있으면 제거
    drop_index_list = []  # 지워버릴 index를 담는 리스트
    for i, row in df.iterrows():
        temp_nouns = row['분석키워드']
        if len(temp_nouns) == 0:  # 만약 명사리스트가 비어 있다면
            drop_index_list.append(i)  # 지울 index 추가
    print(f"drop_index_list: {drop_index_list}")
    df = df.drop(drop_index_list)  # 해당 index를 지우기

    # index를 지우면 순회시 index 값이 중간중간 비기 때문에 index를 다시 지정
    df.index = range(len(df))
    print(df)
    return df


# (기사제목)분석키워드 군집분석
def cluster_analysis(df):

    # 문서를 명사 집합으로 보고 문서 리스트로 치환 (tfidfVectorizer 인풋 형태를 맞추기 위해)
    text = [" ".join(noun) for noun in df['분석키워드']]

    tfidf_vectorizer = TfidfVectorizer(min_df=5, ngram_range=(1, 5))
    tfidf_vectorizer.fit(text)
    vector = tfidf_vectorizer.transform(text).toarray()

    print(vector)

    vector = np.array(vector)  # Normalizer를 이용해 변환된 벡터
    model = DBSCAN(eps=0.3, min_samples=6, metric="cosine")
    # 거리 계산 식으로는 Cosine distance를 이용
    cluster_list = model.fit_predict(vector)

    print(cluster_list)
    df["군집그룹"] = cluster_list

    print(df)

    for cluster_num in set(cluster_list):
        # -1,0은 노이즈 판별이 났거나 클러스터링이 안된 경우
        if (cluster_num == -1 or cluster_num == 0):
            continue
        else:
            print("cluster num : {}".format(cluster_num))
            temp_df = df[df['군집그룹'] == cluster_num]  # cluster num 별로 조회
            for title in temp_df['기사제목']:
                print(title)

    return df

#기사일자 정제화
def news_date_cleansing(news_dates):

    # (A면1단, 날짜) 형태로 되어 있는 경우 처리
    if len(news_dates) > 1:
        news_date = news_dates[1].get_text().replace('.', '')
    else:
        news_date = news_dates[0].get_text().replace('.', '')

    if "분" in news_date or "시간 전" in news_date:
        news_date = str(today_num)
        news_date = datetime.strptime(news_date, '%Y%m%d').strftime('%Y-%m-%d')

    elif "일 전" in news_date:
        news_date_num = int(re.sub('[\D]', '', news_date))
        news_date = str(datetime.today() - timedelta(days=news_date_num))
        news_date = news_date[:10]
    # news_date = datetime.strptime(news_date, '%Y%m%d').strftime('%Y-%m-%d')

    elif "주 전" in news_date:
        news_date_num = int(re.sub('[\D]', '', news_date))
        news_date = str(datetime.today() - timedelta(days=news_date_num * 7))
        news_date = news_date[:10]
    # news_date = datetime.strptime(news_date, '%Y%m%d').strftime('%Y-%m-%d')

    else:
        news_date = str(news_date)
        news_date = datetime.strptime(news_date, '%Y%m%d').strftime('%Y-%m-%d')

    return  re.sub(r'[^0-9]', '', news_date)


#언론사 정제화(ex:"아이뉴스23 언론사 선정" 비예측 패턴 발생)
def press_name_cleansing(press_name):
    # (아이뉴스23 언론사 선정) 형태로 되어 있는 경우 처리
    return press_name.replace('언론사 선정', '')



#본문요약 정제화
def contents_cleansing(contents):
    first_cleansing_contents = re.sub('<dl>.*?</a> </div> </dd> <dd>', '',
                                      str(contents)).strip()  #앞에 필요없는 부분 제거
    second_cleansing_contents = re.sub('<ul class="relation_lst">.*?</dd>', '',
                                       first_cleansing_contents).strip()#뒤에 필요없는 부분 제거 (새끼 기사)
    third_cleansing_contents = re.sub('<.+?>', '', second_cleansing_contents).strip()
    return third_cleansing_contents



def crawler(maxpage, query, sort, s_date, e_date,news_keyword):
    # 각 크롤링 결과 저장하기 위한 리스트 선언
    CRAWL_DATE    = []   # 수집일자
    NEWS_DATE     = []   # 발행일자
    NEWS_KEYWORD  = []   # 검색어
    PRESS_ID      = []   # 언론사ID
    PRESS_NAME    = []   # 언론사
    CLUSTER_GROUP = []   # 군집그룹
    TITLE         = []   # 기사제목
    LINK          = []   # 기사URL
    CONTENT       = []   # 본문요약
    PAGE          = []   # 페이지
    NOUN_LIST     = []   # 분석키워드
    result = {}

    s_from = re.sub(r'[^0-9]', '', s_date)
    e_to   = re.sub(r'[^0-9]', '', e_date)

    page = 1  # 한페이지당 10개 기사
    maxpage_t =(int(maxpage)-1)*10+1   # 11= 2페이지 21=3페이지 31=4페이지  ...81=9페이지 , 91=10페이지, 101=11페이지

    while page <= maxpage_t:
        url = "https://search.naver.com/search.naver?where=news&query=" + query + "&sort="+sort+"&ds=" + s_date + "&de=" + e_date + "&nso=so%3Ar%2Cp%3Afrom" + s_from + "to" + e_to + "%2Ca%3A&start=" + str(page)
        print(url)
        response = requests.get(url)
        time.sleep(1)  # url 웹 페이지 로딩 대기 시간

        html = response.text

        #뷰티풀소프의 인자값 지정
        soup = BeautifulSoup(html, 'html.parser')

        for news_result in soup.select(".list_news > li"):

            news_dates = news_result.select('span.info')
            press_name = news_result.select_one(".info.press").text
            title      = news_result.select_one(".news_tit").text
            link       = news_result.select_one(".news_tit")["href"]
            content    = news_result.select_one(".news_dsc").text

            okt = Okt()  # 형태소 분석기 객체 생성
            nouns = okt.nouns(content)  # 명사만 추출하기, 결과값은 명사 리스트

            CRAWL_DATE.append(today_str)                                            # 수집일자
            NEWS_DATE.append(news_date_cleansing(news_dates))                       # 발행일자
            NEWS_KEYWORD.append(news_keyword)                                       # 검색어
            PRESS_ID.append(PRESS_DIC.get(press_name_cleansing(press_name),'999'))  # 언론사ID
            PRESS_NAME.append(press_name_cleansing(press_name))                     # 언론사
            CLUSTER_GROUP.append("")                                                # 군집그룹
            TITLE.append(title)                                                     # 기사제목
            LINK.append(link)                                                       # 기사링크
            CONTENT.append(contents_cleansing(content))                             # 본문요약
            PAGE.append(page)                                                       # 페이지
            NOUN_LIST.append(nouns)                                                 # 분석키워드

        print(page)
        page += 10


    #모든 리스트 딕셔너리형태로 저장
    result= {
             "수집일자"      : CRAWL_DATE   ,    #A
             "발행일자"      : NEWS_DATE    ,    #B
             "검색어"        : NEWS_KEYWORD ,    #C
             "언론사ID"      : PRESS_ID     ,    #D
             "언론사"        : PRESS_NAME   ,    #E
             "군집그룹"      : CLUSTER_GROUP,    #F
             "기사제목"      : TITLE        ,    #G
             "기사링크"      : LINK         ,    #H
             "본문요약"      : CONTENT      ,    #I
             "페이지"        : PAGE         ,    #J
             "분석키워드"    : NOUN_LIST         #K
             }





    #### 기사수집 메인 df 시작
    df1 = pd.DataFrame(result)  # df로 변환
    df1 = cluster_cleansing(df1)
    df1 = cluster_analysis(df1)

    #df1 정렬
    df1 = df1.sort_values(by=["군집그룹", "발행일자", "언론사"], ascending=[False, False, True])
    #### //기사수집 메인 df 종료



    #### 수집기사 분석키워드 빈도수 집계 df_word_count 시작
    print(df1['분석키워드'])
    list_of_single_column1 = df1['분석키워드'].tolist()           # 분석키워드 dataframe -> (2차원)list 형변환
    list_of_single_column2 = sum(list_of_single_column1,[])      # 분석키워드 (2차원)list -> (1차원)list 형변환
    print(list_of_single_column2)

    # word_count
    word_count = {} # 빈 set
    for noun in list_of_single_column2 :
        word_count[noun] = word_count.get(noun, 0) + 1

    counter = Counter(word_count)
    DESEN_RANK = counter.most_common()
    print(DESEN_RANK)

    df2 = pd.DataFrame(word_count,index = [0])  # df로 변환
    df2 = pd.melt(df2)
    df2.columns = ['키워드', '빈도수']
    df2 = df2.sort_values(by=["빈도수"], ascending=[False])
    print(df2)

    #### // 수집기사 분석키워드 빈도수 집계 df 종료



    ### WORD CLOUD 시작 >> RESULT_FILENAME_CLOUD
    masking_image = np.array(Image.open(".\\img\\mask_bnk_img.png"))
    word_cloud = WordCloud(font_path="C:\\Windows\\Fonts\\HYWULM.TTF", # font_path="C:/Windows/Fonts/NanumSquareB.ttf"
                           width=2000, height=1000,
                           mask=masking_image, # masking
                           colormap='prism', # colormap='autumn',
                           background_color='black',
                           max_font_size=400).generate_from_frequencies(word_count)

    #image_colors = ImageColorGenerator(masking_color)

    RESULT_FILENAME_CLOUD = 'RESULT_%04d%02d%02d_%02d%02d%02d_%s.jpg' % (now.year, now.month, now.day, now.hour, now.minute, now.second,news_keyword)
    word_cloud.to_file(filename = RESULT_PATH + "\\" + RESULT_FILENAME_CLOUD)  # 파일로 저장

    plt.figure(figsize=(10,7))
    plt.imshow(word_cloud, interpolation='bilinear')
    plt.axis("off")
    plt.tight_layout(pad=0)
    #plt.show()
    ### // WORD CLOUD 종료




    # 새로 만들 파일이름 지정 >> RESULT_FILENAME
    RESULT_FILENAME = 'RESULT_%04d%02d%02d_%02d%02d%02d_%s.xlsx' % (now.year, now.month, now.day, now.hour, now.minute, now.second,news_keyword)
    with pd.ExcelWriter(RESULT_PATH + "\\" + RESULT_FILENAME) as writer:
        df1.to_excel(writer, sheet_name  = news_keyword , index = False , freeze_panes= (1,0))  # 틀고정
        df2.to_excel(writer, sheet_name  = "키워드빈도" , index = False , freeze_panes= (1,0))  # 틀고정


    # 엑셀 스타일
    wb = load_workbook(RESULT_PATH + "\\" + RESULT_FILENAME)

    # Sheet1(news_keyword) 스타일
    ws = wb[news_keyword]  # 검색어 시트
    ws.column_dimensions["A"].width = 10.25    # 수집일자
    ws.column_dimensions["B"].width = 10.25    # 발행일자
    ws.column_dimensions["C"].width = 10.25    # 검색어
    ws.column_dimensions["D"].width = 8        # 언론사ID
    ws.column_dimensions["E"].width = 15       # 언론사
    ws.column_dimensions["F"].width = 8        # 군집그룹
    ws.column_dimensions["G"].width = 50       # 기사제목
    ws.column_dimensions["H"].width = 50       # 기사링크
    ws.column_dimensions["I"].width = 100      # 본문요약
    ws.column_dimensions["J"].width = 8        # 페이지
    ws.column_dimensions["K"].width = 50       # 분석키워드


    # 폰트 정의
    font_header  = Font(name="맑은 고딕", size=9, bold=True)
    font_content = Font(name="맑은 고딕", size=9, bold=False)

    cell_alignment_center = Alignment(horizontal    = 'center',  # 좌우 정렬 left, right, center, distributed
                                      vertical      = 'center',  # 위아래 정렬 top, bottom, center, distributed
                                      shrink_to_fit = True,      # 셀의 크기에 맞게 글자를 축소
                                      indent        = 0          # 들여쓰기
                              )

    # CELL 음영 정의
    ligthGrayFill = PatternFill(start_color='00e6e6e6',end_color  ='00e6e6e6',fill_type  ='solid')

    # 헤더CELL 스타일 적용
    for rows in ws["A1":"K1"]:
        for cell in rows:
            cell.font      = font_header
            cell.alignment = cell_alignment_center
            cell.fill      = ligthGrayFill

    # 내용CELL 스타일 적용
    for rows in ws["A2":"K3000"]:
        for cell in rows:
            cell.font      = font_content
            if cell.column_letter in ('A','B','D','F','J') :  #수집일자 / 발행일자 / 언론사ID / 군집그룹 / 페이지
                cell.alignment = cell_alignment_center        # CELL가운데 정렬

    # Sheet3(news_keyword) 스타일
    ws2 = wb.create_sheet("WORD_CLOUD")

    wc_image = Image_openpyxl(RESULT_PATH + "\\" + RESULT_FILENAME_CLOUD)
    ws2.add_image(wc_image,'A1')
    col_width,row_height = wc_image.width,wc_image.height

    ws2.column_dimensions["A"].width = col_width  * 63.2   / 504.19
    ws2.row_dimensions["1"].height   = row_height * 225.35 / 298.96


    # 엑셀 스타일 적용 파일 저장
    wb.save(RESULT_PATH + "\\" + RESULT_FILENAME)
    wb.close()

















def main():

    # 엑셀 파일 열기
    wb = openpyxl.load_workbook(filename = ROOT_PATH + "\\" + "ors_news_crawler.xlsx")

    # 시트 선택하기
    option = wb.get_sheet_by_name('OPTION')

    s_date  = option["A3"].value # 검색 시작일자 YYYY.MM.DD
    e_date  = option["B3"].value # 검색 종료일자 YYYY.MM.DD
    maxpage = option["C3"].value # 검색 page (1page = 기사 10건)
    sort    = option["D3"].value # 검색 옵션(정렬) 관련도순:0, 최신순:1, 오래된순:2

    news_keyword   = ""          # 뉴스 검색 키워드
    query          = ""          # 쿼리 string

    flag = False
    keyword_range = option["A5":"D19"]
    print(keyword_range)
    for keyword in keyword_range:
        query = ""  # 검색어 초기화

        for cell in keyword:
            if cell.column_letter == "A" and cell.value == None:    # 수집포함 메인 keyword None
                flag = True
                break      # 수집중단

            if cell.column_letter == "A" and cell.value != None:  # 수집포함 메인 keyword
                news_keyword = cell.value

            # print("cell_letter:" + str(cell.column_letter))
            # print("keyword:" + str(cell.value))

            if cell.column_letter in ('A','B') and cell.value != None:
                query += " " + "+" + str(cell.value)   # +검색어 생성

            if cell.column_letter in ('C','D') and cell.value != None:
                query += " " + "-" + str(cell.value)   # -검색어 생성

        if flag:
            break #수집 중단

        print("query" + query)
        crawler(maxpage, query, sort, s_date, e_date,news_keyword)


main()
